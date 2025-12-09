import sys
import os
import csv
from datetime import datetime, timedelta
from pathlib import Path

from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                               QHBoxLayout, QPushButton, QTextBrowser, QFileDialog,
                               QLineEdit, QLabel, QGroupBox, QMessageBox, QProgressBar)
from PySide6.QtCore import QThread, Signal, Qt
from PySide6.QtGui import QFont

import pyodbc
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class CDRProcessThread(QThread):
    """CDR 파일 처리를 위한 워커 쓰레드"""
    log_signal = Signal(str)
    progress_signal = Signal(int)
    finished_signal = Signal(bool, str)
    
    def __init__(self, csv_file, server, database, username, password):
        super().__init__()
        self.csv_file = csv_file
        self.server = server
        self.database = database
        self.username = username
        self.password = password
        self.conn = None
        
    def log(self, message):
        """로그 메시지 전송"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_signal.emit(f"[{timestamp}] {message}")
        
    def run(self):
        try:
            # 1. CSV 파일 검증
            self.log("=" * 60)
            self.log("CDR 파일 처리 시작")
            self.log("=" * 60)
            self.progress_signal.emit(5)
            
            if not os.path.exists(self.csv_file):
                raise Exception(f"CSV 파일을 찾을 수 없습니다: {self.csv_file}")
            
            # 파일명에서 날짜 추출 (CDR-25120900.csv -> 25120900)
            filename = Path(self.csv_file).stem
            date_str = filename.split('-')[-1]
            self.log(f"파일명: {filename}")
            self.log(f"추출된 날짜: {date_str}")
            
            # 날짜 파싱 (YYMMDD00 형식)
            try:
                file_date = datetime.strptime(date_str[:6], '%y%m%d')
                # 백업은 다음날 00시에 되므로 하루 전 날짜가 실제 데이터 날짜
                actual_date = file_date - timedelta(days=1)
                formatted_date = actual_date.strftime('%Y%m%d')
                self.log(f"실제 데이터 날짜: {actual_date.strftime('%Y-%m-%d')}")
            except Exception as e:
                raise Exception(f"날짜 파싱 실패: {e}")
            
            self.progress_signal.emit(10)
            
            # 2. CSV 데이터 읽기
            self.log("\nCSV 파일 읽기 중...")
            csv_data = []
            try:
                with open(self.csv_file, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    csv_data = list(reader)
                self.log(f"총 {len(csv_data)}개의 레코드를 읽었습니다.")
            except Exception as e:
                raise Exception(f"CSV 파일 읽기 실패: {e}")
            
            if len(csv_data) == 0:
                raise Exception("CSV 파일에 데이터가 없습니다.")
            
            self.progress_signal.emit(20)
            
            # 3. DB 연결
            self.log("\nSQL Server에 연결 중...")
            try:
                conn_str = (
                    f"DRIVER={{SQL Server}};"
                    f"SERVER={self.server};"
                    f"DATABASE={self.database};"
                    f"UID={self.username};"
                    f"PWD={self.password}"
                )
                self.conn = pyodbc.connect(conn_str)
                self.log("데이터베이스 연결 성공")
            except Exception as e:
                raise Exception(f"DB 연결 실패: {e}")
            
            self.progress_signal.emit(25)
            
            # 4. 임시 테이블 생성
            table_name = f"[{filename}]"
            self.log(f"\n임시 테이블 생성 중: {table_name}")
            cursor = self.conn.cursor()
            
            try:
                # 테이블이 이미 존재하면 삭제
                cursor.execute(f"""
                    IF OBJECT_ID(N'{table_name}', N'U') IS NOT NULL
                        DROP TABLE {table_name}
                """)
                self.conn.commit()
                
                # 테이블 생성
                create_table_sql = f"""
                CREATE TABLE {table_name}(
                    [RecDT] [datetime2](7) NULL,
                    [SendNum] [nvarchar](50) NULL,
                    [RecvNum] [nvarchar](50) NULL,
                    [Gubun] [nvarchar](50) NULL,
                    [StartDT] [datetime2](7) NULL,
                    [EndDT] [datetime2](7) NULL,
                    [CallGubun] [nvarchar](50) NULL,
                    [Result] [nvarchar](50) NULL
                ) ON [PRIMARY]
                """
                cursor.execute(create_table_sql)
                self.conn.commit()
                self.log("테이블 생성 완료")
            except Exception as e:
                raise Exception(f"테이블 생성 실패: {e}")
            
            self.progress_signal.emit(30)
            
            # 5. 데이터 삽입
            self.log(f"\n데이터 삽입 중... (총 {len(csv_data)}개)")
            try:
                insert_sql = f"""
                    INSERT INTO {table_name} 
                    (RecDT, SendNum, RecvNum, Gubun, StartDT, EndDT, CallGubun, Result)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """
                
                batch_size = 1000
                for i in range(0, len(csv_data), batch_size):
                    batch = csv_data[i:i+batch_size]
                    cursor.executemany(insert_sql, batch)
                    self.conn.commit()
                    progress = 30 + int((i / len(csv_data)) * 20)
                    self.progress_signal.emit(progress)
                    if i % 5000 == 0 and i > 0:
                        self.log(f"  {i}개 레코드 삽입 완료...")
                
                self.log(f"전체 데이터 삽입 완료: {len(csv_data)}개")
            except Exception as e:
                raise Exception(f"데이터 삽입 실패: {e}")
            
            self.progress_signal.emit(50)
            
            # 6. 쿼리 실행
            self.log("\n미통화 리스트 조회 중...")
            try:
                query_sql = f"""
                SELECT DISTINCT c1.SendNum AS 발신번호, c2.CntNum AS 통화시도횟수, 
                       ISNULL(s1.SaName,'') AS 담당자, ISNULL(m1.Name,'') AS 성명, 
                       Result AS 통화결과 
                FROM {table_name} c1 WITH(NOLOCK)
                LEFT JOIN dbo.Member m1 WITH(NOLOCK) ON c1.SendNum = REPLACE(m1.Mobile,'-','')
                LEFT JOIN dbo.Staff s1 WITH(NOLOCK) ON m1.Charge_IDP = s1.SaBun
                LEFT JOIN (SELECT SendNum, COUNT(SendNum) AS CntNum FROM {table_name} WITH(NOLOCK) GROUP BY SendNum) c2 
                    ON c1.SendNum = c2.SendNum
                WHERE LEN(c1.SendNum) > 10 AND c1.SendNum NOT IN
                (
                    SELECT SendNum FROM {table_name} WHERE Result = 'Success'
                    UNION ALL
                    SELECT SendNum FROM {table_name} WHERE Result = 'Success'
                    UNION ALL
                    SELECT RecvNum FROM {table_name} WHERE Result = 'Success'
                    UNION ALL
                    SELECT RecvNum FROM {table_name} WHERE Result = 'Success'
                )
                AND CONVERT(CHAR(8),c1.RecDT,8) >= '09:30:00' AND CONVERT(CHAR(8),c1.RecDT,8) < '18:00:00'
                ORDER BY 통화시도횟수 DESC
                """
                
                cursor.execute(query_sql)
                results = cursor.fetchall()
                columns = [column[0] for column in cursor.description]
                self.log(f"미통화 리스트 조회 완료: {len(results)}건")
            except Exception as e:
                raise Exception(f"쿼리 실행 실패: {e}")
            
            self.progress_signal.emit(60)
            
            # 7. 엑셀 파일 생성
            excel_filename = f"{formatted_date}_미통화리스트.xlsx"
            excel_path = os.path.join(os.path.dirname(self.csv_file), excel_filename)
            self.log(f"\n엑셀 파일 생성 중: {excel_filename}")
            
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "미통화리스트"
                
                # 헤더 스타일
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 헤더 작성
                for col_idx, column_name in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=column_name)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                # 데이터 작성
                for row_idx, row_data in enumerate(results, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 열 너비 자동 조정
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(excel_path)
                self.log(f"엑셀 파일 저장 완료: {excel_path}")
            except Exception as e:
                raise Exception(f"엑셀 파일 생성 실패: {e}")
            
            self.progress_signal.emit(75)
            
            # 8. CDR 테이블에 데이터 병합
            self.log("\nCDR 메인 테이블에 데이터 병합 중...")
            try:
                insert_main_sql = f"INSERT INTO CDR SELECT * FROM {table_name}"
                cursor.execute(insert_main_sql)
                affected_rows = cursor.rowcount
                self.conn.commit()
                self.log(f"CDR 테이블에 {affected_rows}개 레코드 추가 완료")
            except Exception as e:
                raise Exception(f"메인 테이블 병합 실패: {e}")
            
            self.progress_signal.emit(90)
            
            # 9. 임시 테이블 삭제
            self.log(f"\n임시 테이블 삭제 중: {table_name}")
            try:
                cursor.execute(f"DROP TABLE {table_name}")
                self.conn.commit()
                self.log("임시 테이블 삭제 완료")
            except Exception as e:
                self.log(f"⚠ 임시 테이블 삭제 경고: {e}")
            
            self.progress_signal.emit(100)
            
            # 완료
            self.log("\n" + "=" * 60)
            self.log("✓ 모든 작업이 성공적으로 완료되었습니다!")
            self.log(f"✓ 엑셀 파일: {excel_path}")
            self.log(f"✓ 미통화 건수: {len(results)}건")
            self.log("=" * 60)
            
            self.finished_signal.emit(True, excel_path)
            
        except Exception as e:
            self.log(f"\n❌ 오류 발생: {str(e)}")
            self.log("=" * 60)
            self.finished_signal.emit(False, str(e))
            
        finally:
            if self.conn:
                self.conn.close()
                self.log("\n데이터베이스 연결 종료")


class CDRProcessorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.thread = None
        self.init_ui()
        
    def init_ui(self):
        self.setWindowTitle("CDR 파일 처리 프로그램 v1.0")
        self.setGeometry(100, 100, 900, 700)
        
        # 중앙 위젯
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(15, 15, 15, 15)
        
        # 1. 파일 선택 그룹
        file_group = QGroupBox("1. CDR 파일 선택")
        file_group.setFont(QFont("맑은 고딕", 10, QFont.Bold))
        file_layout = QHBoxLayout()
        
        self.file_path_edit = QLineEdit()
        self.file_path_edit.setPlaceholderText("CDR CSV 파일을 선택하세요 (예: CDR-25120900.csv)")
        self.file_path_edit.setReadOnly(True)
        
        self.file_btn = QPushButton("파일 선택")
        self.file_btn.setFixedWidth(100)
        self.file_btn.clicked.connect(self.select_file)
        
        file_layout.addWidget(self.file_path_edit)
        file_layout.addWidget(self.file_btn)
        file_group.setLayout(file_layout)
        
        # 2. DB 연결 정보 그룹
        db_group = QGroupBox("2. 데이터베이스 연결 정보")
        db_group.setFont(QFont("맑은 고딕", 10, QFont.Bold))
        db_layout = QVBoxLayout()
        
        # 서버
        server_layout = QHBoxLayout()
        server_layout.addWidget(QLabel("서버:"))
        self.server_edit = QLineEdit()
        self.server_edit.setPlaceholderText("localhost 또는 IP주소")
        server_layout.addWidget(self.server_edit)
        
        # 데이터베이스
        db_name_layout = QHBoxLayout()
        db_name_layout.addWidget(QLabel("데이터베이스:"))
        self.database_edit = QLineEdit()
        self.database_edit.setPlaceholderText("데이터베이스 이름")
        db_name_layout.addWidget(self.database_edit)
        
        # 사용자명
        user_layout = QHBoxLayout()
        user_layout.addWidget(QLabel("사용자명:"))
        self.username_edit = QLineEdit()
        self.username_edit.setPlaceholderText("SQL Server 사용자명")
        user_layout.addWidget(self.username_edit)
        
        # 비밀번호
        pass_layout = QHBoxLayout()
        pass_layout.addWidget(QLabel("비밀번호:"))
        self.password_edit = QLineEdit()
        self.password_edit.setEchoMode(QLineEdit.Password)
        self.password_edit.setPlaceholderText("SQL Server 비밀번호")
        pass_layout.addWidget(self.password_edit)
        
        db_layout.addLayout(server_layout)
        db_layout.addLayout(db_name_layout)
        db_layout.addLayout(user_layout)
        db_layout.addLayout(pass_layout)
        db_group.setLayout(db_layout)
        
        # 3. 진행 상태 그룹
        progress_group = QGroupBox("3. 처리 진행 상태")
        progress_group.setFont(QFont("맑은 고딕", 10, QFont.Bold))
        progress_layout = QVBoxLayout()
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        
        self.log_browser = QTextBrowser()
        self.log_browser.setFont(QFont("Consolas", 9))
        
        progress_layout.addWidget(self.progress_bar)
        progress_layout.addWidget(self.log_browser)
        progress_group.setLayout(progress_layout)
        
        # 4. 실행 버튼
        button_layout = QHBoxLayout()
        
        self.start_btn = QPushButton("처리 시작")
        self.start_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 14px;
                font-weight: bold;
                padding: 10px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)
        self.start_btn.clicked.connect(self.start_process)
        
        self.clear_btn = QPushButton("로그 지우기")
        self.clear_btn.setFixedWidth(120)
        self.clear_btn.clicked.connect(self.clear_log)
        
        button_layout.addWidget(self.start_btn)
        button_layout.addWidget(self.clear_btn)
        
        # 레이아웃 구성
        main_layout.addWidget(file_group)
        main_layout.addWidget(db_group)
        main_layout.addWidget(progress_group, 1)
        main_layout.addLayout(button_layout)
        
        # 초기 로그
        self.log_browser.append("CDR 파일 처리 프로그램을 시작합니다.")
        self.log_browser.append("1. CDR CSV 파일을 선택하세요.")
        self.log_browser.append("2. 데이터베이스 연결 정보를 입력하세요.")
        self.log_browser.append("3. '처리 시작' 버튼을 클릭하세요.")
        self.log_browser.append("-" * 60)
        
    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "CDR CSV 파일 선택",
            "",
            "CSV Files (*.csv);;All Files (*)"
        )
        if file_path:
            self.file_path_edit.setText(file_path)
            self.log_browser.append(f"✓ 파일 선택됨: {os.path.basename(file_path)}")
            
    def clear_log(self):
        self.log_browser.clear()
        self.progress_bar.setValue(0)
        
    def validate_inputs(self):
        """입력값 검증"""
        if not self.file_path_edit.text():
            QMessageBox.warning(self, "입력 오류", "CDR 파일을 선택해주세요.")
            return False
            
        if not self.server_edit.text():
            QMessageBox.warning(self, "입력 오류", "서버 주소를 입력해주세요.")
            return False
            
        if not self.database_edit.text():
            QMessageBox.warning(self, "입력 오류", "데이터베이스 이름을 입력해주세요.")
            return False
            
        if not self.username_edit.text():
            QMessageBox.warning(self, "입력 오류", "사용자명을 입력해주세요.")
            return False
            
        if not self.password_edit.text():
            QMessageBox.warning(self, "입력 오류", "비밀번호를 입력해주세요.")
            return False
            
        return True
        
    def start_process(self):
        """처리 시작"""
        if not self.validate_inputs():
            return
            
        # UI 비활성화
        self.start_btn.setEnabled(False)
        self.file_btn.setEnabled(False)
        self.progress_bar.setValue(0)
        self.clear_log()
        
        # 워커 쓰레드 생성 및 시작
        self.thread = CDRProcessThread(
            self.file_path_edit.text(),
            self.server_edit.text(),
            self.database_edit.text(),
            self.username_edit.text(),
            self.password_edit.text()
        )
        
        self.thread.log_signal.connect(self.update_log)
        self.thread.progress_signal.connect(self.update_progress)
        self.thread.finished_signal.connect(self.process_finished)
        
        self.thread.start()
        
    def update_log(self, message):
        """로그 업데이트"""
        self.log_browser.append(message)
        self.log_browser.verticalScrollBar().setValue(
            self.log_browser.verticalScrollBar().maximum()
        )
        
    def update_progress(self, value):
        """진행률 업데이트"""
        self.progress_bar.setValue(value)
        
    def process_finished(self, success, result):
        """처리 완료"""
        self.start_btn.setEnabled(True)
        self.file_btn.setEnabled(True)
        
        if success:
            QMessageBox.information(
                self,
                "처리 완료",
                f"CDR 파일 처리가 완료되었습니다.\n\n생성된 파일:\n{result}"
            )
        else:
            QMessageBox.critical(
                self,
                "처리 실패",
                f"오류가 발생했습니다:\n\n{result}\n\n로그를 확인해주세요."
            )


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    window = CDRProcessorApp()
    window.show()
    
    sys.exit(app.exec())


if __name__ == '__main__':
    main()