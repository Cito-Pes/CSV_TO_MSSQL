import sys
import os
import csv
import re
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path

from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                               QHBoxLayout, QPushButton, QTextBrowser, QFileDialog,
                               QLineEdit, QLabel, QGroupBox, QMessageBox, QProgressBar)
from PySide6.QtCore import QThread, Signal, Qt
from PySide6.QtGui import QFont, QIcon

import requests
import pyodbc
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def download_db():
    """구글 드라이브에서 Config_DB.db 파일 다운로드"""
    url = "https://drive.google.com/file/d/1oncya1uYDnbVS2KwuBAKw4x4o9oQDct0/view?usp=drive_link"
    db_dir = "./DB"
    db_path = os.path.join(db_dir, "Config_DB.db")
    
    if not os.path.exists(db_dir):
        os.makedirs(db_dir)
    
    try:
        # 파일 id 추출
        match = re.search(r"/d/([a-zA-Z0-9_-]+)", url)
        if not match:
            raise Exception("구글 드라이브 파일 ID를 찾을 수 없습니다.")
        file_id = match.group(1)
        download_url = f"https://drive.google.com/uc?export=download&id={file_id}"

        session = requests.Session()
        resp = session.get(download_url, stream=True)
        resp.raise_for_status()

        # 구글 드라이브는 큰 파일일 경우 'confirm' 토큰이 필요함
        if "text/html" in resp.headers.get("Content-Type", ""):
            # 토큰 추출
            for key, value in resp.cookies.items():
                if key.startswith("download_warning"):
                    confirm_token = value
                    download_url = f"https://drive.google.com/uc?export=download&confirm={confirm_token}&id={file_id}"
                    resp = session.get(download_url, stream=True)
                    resp.raise_for_status()
                    break

        with open(db_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
        return True, db_path
    except Exception as e:
        return False, str(e)


def load_db_config():
    """Config_DB.db에서 DB 연결 정보 로드"""
    db_path = os.path.join("./DB", "Config_DB.db")
    
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        query = "SELECT DB_Type, Host, Port, DB_Name, DB_ID, DB_PW FROM DBCON WHERE Name = 'HD_MSSQL'"
        cursor.execute(query)
        result = cursor.fetchone()
        conn.close()
        
        if result:
            return {
                'DB_Type': result[0],
                'Host': result[1],
                'Port': result[2],
                'DB_Name': result[3],
                'DB_ID': result[4],
                'DB_PW': result[5]
            }
        else:
            raise Exception("HD_MSSQL 설정을 찾을 수 없습니다.")
            
    except Exception as e:
        raise Exception(f"DB 설정 로드 실패: {e}")


class CDRProcessThread(QThread):
    """CDR 파일 처리를 위한 워커 쓰레드"""
    log_signal = Signal(str)
    progress_signal = Signal(int)
    finished_signal = Signal(bool, str)
    
    def __init__(self, csv_file, db_config):
        super().__init__()
        self.csv_file = csv_file
        self.db_config = db_config
        self.conn = None
        
    def log(self, message):
        """로그 메시지 전송"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_signal.emit(f"[{timestamp}] {message}")
        
    def run(self):
        try:
            # 1. CSV 파일 검증
            self.log("=" * 60)
            self.log("CDR 파일 처리 시작")
            self.log("=" * 60)
            self.progress_signal.emit(5)
            
            if not os.path.exists(self.csv_file):
                raise Exception(f"CSV 파일을 찾을 수 없습니다: {self.csv_file}")
            
            # 파일명에서 날짜 추출 (CDR-25120900.csv -> 25120900)
            filename = Path(self.csv_file).stem
            date_str = filename.split('-')[-1]
            self.log(f"파일명: {filename}")
            self.log(f"추출된 날짜: {date_str}")
            
            # 날짜 파싱 (YYMMDD00 형식)
            try:
                file_date = datetime.strptime(date_str[:6], '%y%m%d')
                # 백업은 다음날 00시에 되므로 하루 전 날짜가 실제 데이터 날짜
                actual_date = file_date - timedelta(days=1)
                formatted_date = actual_date.strftime('%Y%m%d')
                self.log(f"실제 데이터 날짜: {actual_date.strftime('%Y-%m-%d')}")
            except Exception as e:
                raise Exception(f"날짜 파싱 실패: {e}")
            
            self.progress_signal.emit(10)
            
            # 2. CSV 데이터 읽기
            self.log("\nCSV 파일 읽기 중...")
            csv_data = []
            try:
                with open(self.csv_file, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    csv_data = list(reader)
                self.log(f"총 {len(csv_data)}개의 레코드를 읽었습니다.")
            except Exception as e:
                raise Exception(f"CSV 파일 읽기 실패: {e}")
            
            if len(csv_data) == 0:
                raise Exception("CSV 파일에 데이터가 없습니다.")
            
            self.progress_signal.emit(20)
            
            # 3. DB 연결
            self.log("\nSQL Server에 연결 중...")
            self.log(f"서버: {self.db_config['Host']}:{self.db_config['Port']}")
            self.log(f"데이터베이스: {self.db_config['DB_Name']}")
            
            try:
                conn_str = (
                    f"DRIVER={{{self.db_config['DB_Type']}}};"
                    f"SERVER={self.db_config['Host']},{self.db_config['Port']};"
                    f"DATABASE={self.db_config['DB_Name']};"
                    f"UID={self.db_config['DB_ID']};"
                    f"PWD={self.db_config['DB_PW']}"
                )
                self.conn = pyodbc.connect(conn_str)
                self.log("데이터베이스 연결 성공")
            except Exception as e:
                raise Exception(f"DB 연결 실패: {e}")
            
            self.progress_signal.emit(25)
            
            # 4. 임시 테이블 생성
            table_name = f"[{filename}]"
            self.log(f"\n임시 테이블 생성 중: {table_name}")
            cursor = self.conn.cursor()
            
            try:
                # 테이블이 이미 존재하면 삭제
                cursor.execute(f"""
                    IF OBJECT_ID(N'{table_name}', N'U') IS NOT NULL
                        DROP TABLE {table_name}
                """)
                self.conn.commit()
                
                # 테이블 생성
                create_table_sql = f"""
                CREATE TABLE {table_name}(
                    [RecDT] [datetime2](7) NULL,
                    [SendNum] [nvarchar](50) NULL,
                    [RecvNum] [nvarchar](50) NULL,
                    [Gubun] [nvarchar](50) NULL,
                    [StartDT] [datetime2](7) NULL,
                    [EndDT] [datetime2](7) NULL,
                    [CallGubun] [nvarchar](50) NULL,
                    [Result] [nvarchar](50) NULL
                ) ON [PRIMARY]
                """
                cursor.execute(create_table_sql)
                self.conn.commit()
                self.log("테이블 생성 완료")
            except Exception as e:
                raise Exception(f"테이블 생성 실패: {e}")
            
            self.progress_signal.emit(30)
            
            # 5. 데이터 삽입
            self.log(f"\n데이터 삽입 중... (총 {len(csv_data)}개)")
            try:
                insert_sql = f"""
                    INSERT INTO {table_name} 
                    (RecDT, SendNum, RecvNum, Gubun, StartDT, EndDT, CallGubun, Result)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """
                
                # 데이터 전처리: 빈 문자열을 None으로 변환
                processed_data = []
                for row in csv_data:
                    processed_row = []
                    for value in row:
                        # 빈 문자열이나 공백만 있는 경우 None으로 변환
                        if value is None or (isinstance(value, str) and value.strip() == ''):
                            processed_row.append(None)
                        else:
                            processed_row.append(value)
                    processed_data.append(tuple(processed_row))
                
                batch_size = 1000
                for i in range(0, len(processed_data), batch_size):
                    batch = processed_data[i:i+batch_size]
                    cursor.executemany(insert_sql, batch)
                    self.conn.commit()
                    progress = 30 + int((i / len(processed_data)) * 20)
                    self.progress_signal.emit(progress)
                    if i % 5000 == 0 and i > 0:
                        self.log(f"  {i}개 레코드 삽입 완료...")
                
                self.log(f"전체 데이터 삽입 완료: {len(processed_data)}개")
            except Exception as e:
                raise Exception(f"데이터 삽입 실패: {e}")
            
            self.progress_signal.emit(50)
            
            # 6. 쿼리 실행
            self.log("\n미통화 리스트 조회 중...")
            try:
                query_sql = f"""
                SELECT DISTINCT c1.SendNum AS 발신번호, c2.CntNum AS 통화시도횟수, 
                       ISNULL(s1.SaName,'') AS 담당자, ISNULL(m1.Name,'') AS 성명, 
                       Result AS 통화결과 
                FROM {table_name} c1 WITH(NOLOCK)
                LEFT JOIN dbo.Member m1 WITH(NOLOCK) ON c1.SendNum = REPLACE(m1.Mobile,'-','')
                LEFT JOIN dbo.Staff s1 WITH(NOLOCK) ON m1.Charge_IDP = s1.SaBun
                LEFT JOIN (SELECT SendNum, COUNT(SendNum) AS CntNum FROM {table_name} WITH(NOLOCK) GROUP BY SendNum) c2 
                    ON c1.SendNum = c2.SendNum
                WHERE LEN(c1.SendNum) > 10 AND c1.SendNum NOT IN
                (
                    SELECT SendNum FROM {table_name} WHERE Result = 'Success'
                    UNION ALL
                    SELECT SendNum FROM {table_name} WHERE Result = 'Success'
                    UNION ALL
                    SELECT RecvNum FROM {table_name} WHERE Result = 'Success'
                    UNION ALL
                    SELECT RecvNum FROM {table_name} WHERE Result = 'Success'
                )
                AND CONVERT(CHAR(8),c1.RecDT,8) >= '09:30:00' AND CONVERT(CHAR(8),c1.RecDT,8) < '18:00:00'
                ORDER BY 통화시도횟수 DESC
                """
                
                cursor.execute(query_sql)
                results = cursor.fetchall()
                columns = [column[0] for column in cursor.description]
                self.log(f"미통화 리스트 조회 완료: {len(results)}건")
            except Exception as e:
                raise Exception(f"쿼리 실행 실패: {e}")
            
            self.progress_signal.emit(60)
            
            # 7. 엑셀 파일 생성
            excel_filename = f"{formatted_date}_미통화리스트.xlsx"
            excel_path = os.path.join(os.path.dirname(self.csv_file), excel_filename)
            self.log(f"\n엑셀 파일 생성 중: {excel_filename}")
            
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "미통화리스트"
                
                # 헤더 스타일
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 헤더 작성
                for col_idx, column_name in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=column_name)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                # 데이터 작성
                for row_idx, row_data in enumerate(results, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 열 너비 자동 조정
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(excel_path)
                self.log(f"엑셀 파일 저장 완료: {excel_path}")
            except Exception as e:
                raise Exception(f"엑셀 파일 생성 실패: {e}")
            
            self.progress_signal.emit(75)
            
            # 8. CDR 테이블에 데이터 병합
            self.log("\nCDR 메인 테이블에 데이터 병합 중...")
            try:
                insert_main_sql = f"INSERT INTO CDR SELECT * FROM {table_name}"
                cursor.execute(insert_main_sql)
                affected_rows = cursor.rowcount
                self.conn.commit()
                self.log(f"CDR 테이블에 {affected_rows}개 레코드 추가 완료")
            except Exception as e:
                raise Exception(f"메인 테이블 병합 실패: {e}")
            
            self.progress_signal.emit(90)
            
            # 9. 임시 테이블 삭제
            self.log(f"\n임시 테이블 삭제 중: {table_name}")
            try:
                cursor.execute(f"DROP TABLE {table_name}")
                self.conn.commit()
                self.log("임시 테이블 삭제 완료")
            except Exception as e:
                self.log(f"⚠ 임시 테이블 삭제 경고: {e}")
            
            self.progress_signal.emit(100)
            
            # 완료
            self.log("\n" + "=" * 60)
            self.log("✓ 모든 작업이 성공적으로 완료되었습니다!")
            self.log(f"✓ 엑셀 파일: {excel_path}")
            self.log(f"✓ 미통화 건수: {len(results)}건")
            self.log("=" * 60)
            
            self.finished_signal.emit(True, excel_path)
            
        except Exception as e:
            self.log(f"\n❌ 오류 발생: {str(e)}")
            self.log("=" * 60)
            self.finished_signal.emit(False, str(e))
            
        finally:
            if self.conn:
                self.conn.close()
                self.log("\n데이터베이스 연결 종료")


class CDRProcessorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.thread = None
        self.db_config = None
        self.init_ui()
        self.load_config()
        
    def init_ui(self):
        self.setWindowTitle("CDR 파일 처리 프로그램 v2.0")
        self.setGeometry(100, 100, 900, 700)
        
        # 프로그램 아이콘 설정
        icon_path = "images/icon.ico"
        if os.path.exists(icon_path):
            from PySide6.QtGui import QIcon
            self.setWindowIcon(QIcon(icon_path))
        
        # 중앙 위젯
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(15, 15, 15, 15)
        
        # 1. 파일 선택 그룹
        file_group = QGroupBox("1. CDR 파일 선택")
        file_group.setFont(QFont("맑은 고딕", 10, QFont.Bold))
        file_layout = QHBoxLayout()
        
        self.file_path_edit = QLineEdit()
        self.file_path_edit.setPlaceholderText("CDR CSV 파일을 선택하세요 (예: CDR-25120900.csv)")
        self.file_path_edit.setReadOnly(True)
        
        self.file_btn = QPushButton("파일 선택")
        self.file_btn.setFixedWidth(100)
        self.file_btn.clicked.connect(self.select_file)
        
        file_layout.addWidget(self.file_path_edit)
        file_layout.addWidget(self.file_btn)
        file_group.setLayout(file_layout)
        
        # 2. DB 연결 정보 그룹 (읽기 전용)
        db_group = QGroupBox("2. 데이터베이스 연결 정보 (자동 로드)")
        db_group.setFont(QFont("맑은 고딕", 10, QFont.Bold))
        db_layout = QVBoxLayout()
        
        # 서버
        server_layout = QHBoxLayout()
        server_layout.addWidget(QLabel("서버:"))
        self.server_edit = QLineEdit()
        self.server_edit.setReadOnly(True)
        self.server_edit.setStyleSheet("background-color: #f0f0f0;")
        server_layout.addWidget(self.server_edit)
        
        # 포트
        port_layout = QHBoxLayout()
        port_layout.addWidget(QLabel("포트:"))
        self.port_edit = QLineEdit()
        self.port_edit.setReadOnly(True)
        self.port_edit.setStyleSheet("background-color: #f0f0f0;")
        port_layout.addWidget(self.port_edit)
        
        # 데이터베이스
        db_name_layout = QHBoxLayout()
        db_name_layout.addWidget(QLabel("데이터베이스:"))
        self.database_edit = QLineEdit()
        self.database_edit.setReadOnly(True)
        self.database_edit.setStyleSheet("background-color: #f0f0f0;")
        db_name_layout.addWidget(self.database_edit)
        
        # 사용자명 (마스킹)
        user_layout = QHBoxLayout()
        user_layout.addWidget(QLabel("사용자명:"))
        self.username_edit = QLineEdit()
        self.username_edit.setReadOnly(True)
        self.username_edit.setStyleSheet("background-color: #f0f0f0;")
        user_layout.addWidget(self.username_edit)
        
        # 비밀번호 (마스킹)
        pass_layout = QHBoxLayout()
        pass_layout.addWidget(QLabel("비밀번호:"))
        self.password_edit = QLineEdit()
        self.password_edit.setReadOnly(True)
        self.password_edit.setEchoMode(QLineEdit.Password)
        self.password_edit.setStyleSheet("background-color: #f0f0f0;")
        pass_layout.addWidget(self.password_edit)
        
        db_layout.addLayout(server_layout)
        db_layout.addLayout(port_layout)
        db_layout.addLayout(db_name_layout)
        db_layout.addLayout(user_layout)
        db_layout.addLayout(pass_layout)
        db_group.setLayout(db_layout)
        
        # 3. 진행 상태 그룹
        progress_group = QGroupBox("3. 처리 진행 상태")
        progress_group.setFont(QFont("맑은 고딕", 10, QFont.Bold))
        progress_layout = QVBoxLayout()
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        
        self.log_browser = QTextBrowser()
        self.log_browser.setFont(QFont("Consolas", 9))
        
        progress_layout.addWidget(self.progress_bar)
        progress_layout.addWidget(self.log_browser)
        progress_group.setLayout(progress_layout)
        
        # 4. 실행 버튼
        button_layout = QHBoxLayout()
        
        self.start_btn = QPushButton("처리 시작")
        self.start_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 14px;
                font-weight: bold;
                padding: 10px;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)
        self.start_btn.clicked.connect(self.start_process)
        self.start_btn.setEnabled(False)  # 초기에는 비활성화
        
        self.clear_btn = QPushButton("로그 지우기")
        self.clear_btn.setFixedWidth(120)
        self.clear_btn.clicked.connect(self.clear_log)
        
        button_layout.addWidget(self.start_btn)
        button_layout.addWidget(self.clear_btn)
        
        # 레이아웃 구성
        main_layout.addWidget(file_group)
        main_layout.addWidget(db_group)
        main_layout.addWidget(progress_group, 1)
        main_layout.addLayout(button_layout)
        
        # 초기 로그
        self.log_browser.append("=" * 60)
        self.log_browser.append("CDR 파일 처리 프로그램 v2.0")
        self.log_browser.append("=" * 60)
        self.log_browser.append("프로그램 초기화 중...")
        self.log_browser.append("-" * 60)
        
    def load_config(self):
        """Config DB 로드"""
        self.log_browser.append("\n데이터베이스 설정 파일 확인 중...")
        
        db_path = os.path.join("./DB", "Config_DB.db")
        
        # DB 파일이 없으면 다운로드
        if not os.path.exists(db_path):
            self.log_browser.append("Config_DB.db 파일이 없습니다.")
            self.log_browser.append("구글 드라이브에서 다운로드 중...")
            
            success, result = download_db()
            
            if success:
                self.log_browser.append(f"✓ DB 파일 다운로드 완료: {result}")
            else:
                self.log_browser.append(f"❌ DB 파일 다운로드 실패: {result}")
                QMessageBox.critical(
                    self,
                    "설정 로드 실패",
                    f"Config_DB.db 파일을 다운로드할 수 없습니다.\n\n오류: {result}"
                )
                return
        else:
            self.log_browser.append(f"✓ Config_DB.db 파일 존재: {db_path}")
        
        # DB 설정 로드
        try:
            self.log_browser.append("\nDB 연결 정보 로드 중...")
            self.db_config = load_db_config()
            
            # UI에 정보 표시 (마스킹)
            self.server_edit.setText(self.db_config['Host'])
            self.port_edit.setText(str(self.db_config['Port']))
            self.database_edit.setText(self.db_config['DB_Name'])
            
            # 사용자명과 비밀번호는 마스킹 처리
            username_masked = '*' * len(self.db_config['DB_ID'])
            password_masked = '*' * len(self.db_config['DB_PW'])
            self.username_edit.setText(username_masked)
            self.password_edit.setText(password_masked)
            
            self.log_browser.append(f"✓ DB 타입: {self.db_config['DB_Type']}")
            self.log_browser.append(f"✓ 서버: {self.db_config['Host']}:{self.db_config['Port']}")
            self.log_browser.append(f"✓ 데이터베이스: {self.db_config['DB_Name']}")
            self.log_browser.append(f"✓ 사용자: {username_masked}")
            self.log_browser.append("\n" + "=" * 60)
            self.log_browser.append("✓ 데이터베이스 설정 로드 완료")
            self.log_browser.append("=" * 60)
            self.log_browser.append("\n1. CDR CSV 파일을 선택하세요.")
            self.log_browser.append("2. '처리 시작' 버튼을 클릭하세요.")
            self.log_browser.append("-" * 60)
            
            # 설정이 로드되면 시작 버튼 활성화
            self.start_btn.setEnabled(True)
            
        except Exception as e:
            self.log_browser.append(f"❌ DB 설정 로드 실패: {str(e)}")
            QMessageBox.critical(
                self,
                "설정 로드 실패",
                f"데이터베이스 설정을 로드할 수 없습니다.\n\n오류: {str(e)}"
            )
        
    def select_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "CDR CSV 파일 선택",
            "",
            "CSV Files (*.csv);;All Files (*)"
        )
        if file_path:
            self.file_path_edit.setText(file_path)
            self.log_browser.append(f"\n✓ 파일 선택됨: {os.path.basename(file_path)}")
            
    def clear_log(self):
        self.log_browser.clear()
        self.progress_bar.setValue(0)
        
    def validate_inputs(self):
        """입력값 검증"""
        if not self.file_path_edit.text():
            QMessageBox.warning(self, "입력 오류", "CDR 파일을 선택해주세요.")
            return False
            
        if not self.db_config:
            QMessageBox.warning(self, "설정 오류", "데이터베이스 설정이 로드되지 않았습니다.")
            return False
            
        return True
        
    def start_process(self):
        """처리 시작"""
        if not self.validate_inputs():
            return
            
        # UI 비활성화
        self.start_btn.setEnabled(False)
        self.file_btn.setEnabled(False)
        self.progress_bar.setValue(0)
        self.clear_log()
        
        # 워커 쓰레드 생성 및 시작
        self.thread = CDRProcessThread(
            self.file_path_edit.text(),
            self.db_config
        )
        
        self.thread.log_signal.connect(self.update_log)
        self.thread.progress_signal.connect(self.update_progress)
        self.thread.finished_signal.connect(self.process_finished)
        
        self.thread.start()
        
    def update_log(self, message):
        """로그 업데이트"""
        self.log_browser.append(message)
        self.log_browser.verticalScrollBar().setValue(
            self.log_browser.verticalScrollBar().maximum()
        )
        
    def update_progress(self, value):
        """진행률 업데이트"""
        self.progress_bar.setValue(value)
        
    def process_finished(self, success, result):
        """처리 완료"""
        self.start_btn.setEnabled(True)
        self.file_btn.setEnabled(True)
        
        if success:
            QMessageBox.information(
                self,
                "처리 완료",
                f"CDR 파일 처리가 완료되었습니다.\n\n생성된 파일:\n{result}"
            )
        else:
            QMessageBox.critical(
                self,
                "처리 실패",
                f"오류가 발생했습니다:\n\n{result}\n\n로그를 확인해주세요."
            )


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    window = CDRProcessorApp()
    window.show()
    
    sys.exit(app.exec())


if __name__ == '__main__':
    main()